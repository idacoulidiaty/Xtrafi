from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import plotly.graph_objects as go


def export_excel_with_figures(df_list, fig_list):
    output = BytesIO()
    wb = Workbook()
    del wb["Sheet"]  # Supprimer la feuille par défaut

    # === 1. EXPORT TABLEAUX AVEC STYLE SIMILAIRE À STREAMLIT ===
    for df, sheet_name in df_list:
        ws = wb.create_sheet(title=sheet_name[:31])  # max 31 caractères

        # Ajouter les lignes du DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)

        # Mise en forme : en-tête
        header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Bordures fines + alignement du contenu
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # Ajustement automatique des colonnes
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # === 2. INSERTION DES GRAPHIQUES EN 2 COLONNES ===
    if fig_list:
        ws_graph = wb.create_sheet(title="Graphiques")

        col_positions = ["B", "M"]  # colonnes pour les deux blocs
        x_index = 0
        y_cursor = [2, 2]  # position verticale pour chaque colonne

        for fig in fig_list:
            if fig is None:
                continue

            img_buffer = BytesIO()
            fig.write_image(img_buffer, format="png")
            img_buffer.seek(0)
            img = XLImage(img_buffer)

            # Choisir colonne et ligne
            col_letter = col_positions[x_index % 2]
            row_number = y_cursor[x_index % 2]
            cell_location = f"{col_letter}{row_number}"

            ws_graph.add_image(img, cell_location)

            # Incrémenter la hauteur dans la bonne colonne
            y_cursor[x_index % 2] += 25
            x_index += 1

    # Finalisation
    wb.save(output)
    output.seek(0)
    return output.getvalue()

